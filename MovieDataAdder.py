'''
This code will add the names of the director of the movie. It can be reformatted to add other information such as lead actor and 
writer but you will have to sift through the HTML code of Letterboxd in order to find how to retrieve that information.
'''
import time
import requests
import openpyxl
from bs4 import BeautifulSoup

# The location of your file with the movie data
loc = "filename"
wb = openpyxl.load_workbook(loc)
sheet = wb.active

# This will add a header for the Director column
f = sheet.cell(1, 6)
f.value = "Director"

# This is used to provide updates
runtimes = []
for i in range(2, sheet.max_row + 1):
    start = time.time()
    url = sheet.cell(i, 4).value
    page = requests.get(url)
    soup = BeautifulSoup(page.text, 'html.parser')
    c = sheet.cell(i, 6)

    # This adds the director's name to the column and it changes the commas because we want to import
    # a csv file and they will cause trouble.
    c.value = soup.find("meta", {"name": "twitter:data1"})['content'].replace(',', ' and')
    finish = time.time()
    runtimes.append(finish - start)
    if (i - 1) % 10 == 0:
        print(i - 1, "done", sheet.max_row + 1 - (i - 1), "left",
              int((sum(runtimes)/len(runtimes)) * (sheet.max_row + 1 - (i - 1)) / 60) + 1, "minutes left")

wb.save("filename")
